from __future__ import annotations

import asyncio
import io
import json
import logging
from dataclasses import asdict
from datetime import datetime
from typing import Any, Dict, List, Optional, Tuple
from zoneinfo import ZoneInfo

from aiogram.exceptions import TelegramBadRequest
from aiogram.types import CallbackQuery, InlineKeyboardButton, InlineKeyboardMarkup
from aiogram.utils.keyboard import InlineKeyboardBuilder
from openpyxl import Workbook, load_workbook
from openpyxl.worksheet.worksheet import Worksheet
from webdav3.client import Client

from config import (
    COMPANIES,
    STATUS_APPROVED,
    STATUS_ISSUED,
    STATUS_REJECTED,
    STATUS_REQUESTED,
    STATUS_RETURNED,
)

log = logging.getLogger(__name__)

# -------------------------
# Callback data helpers
# -------------------------
CB_PREFIX = "act"


def pack_cb(action: str, value: str) -> str:
    # value должен быть короткий (TG лимит на callback_data)
    return f"{CB_PREFIX}:{action}:{value}"


def unpack_cb(data: str) -> Tuple[str, str]:
    parts = (data or "").split(":", 2)
    if len(parts) != 3 or parts[0] != CB_PREFIX:
        raise ValueError("Bad callback data")
    return parts[1], parts[2]


# -------------------------
# Keyboards
# -------------------------
def kb_companies() -> InlineKeyboardMarkup:
    builder = InlineKeyboardBuilder()
    for idx, name in enumerate(COMPANIES):
        builder.add(InlineKeyboardButton(text=name, callback_data=pack_cb("cmp", str(idx))))
    builder.adjust(1)
    return builder.as_markup()


def kb_tokens(tokens: List[Dict[str, Any]]) -> InlineKeyboardMarkup:
    builder = InlineKeyboardBuilder()
    for t in tokens:
        title = t["token_id"]
        if t.get("description"):
            title = f'{t["token_id"]} — {t["description"]}'
        builder.add(InlineKeyboardButton(text=title, callback_data=pack_cb("tok", t["token_id"])))
    builder.adjust(1)
    return builder.as_markup()


def kb_director_decision(request_id: int) -> InlineKeyboardMarkup:
    builder = InlineKeyboardBuilder()
    builder.add(
        InlineKeyboardButton(text="✅ Одобрить", callback_data=pack_cb("apr", str(request_id))),
        InlineKeyboardButton(text="❌ Отклонить", callback_data=pack_cb("rej", str(request_id))),
    )
    builder.adjust(2)
    return builder.as_markup()


def kb_officer_actions(request_id: int, status: str) -> InlineKeyboardMarkup:
    builder = InlineKeyboardBuilder()
    if status == STATUS_APPROVED:
        builder.add(InlineKeyboardButton(text="📦 Выдал (все токены)", callback_data=pack_cb("iss", str(request_id))))
    elif status == STATUS_ISSUED:
        builder.add(InlineKeyboardButton(text="✅ Принял (все токены)", callback_data=pack_cb("ret", str(request_id))))
    builder.adjust(1)
    return builder.as_markup()


# -------------------------
# Text formatters
# -------------------------
def escape_html(text: Optional[str]) -> str:
    if text is None:
        return ""
    return (
        str(text)
        .replace("&", "&amp;")
        .replace("<", "&lt;")
        .replace(">", "&gt;")
    )


def status_human(status: str) -> str:
    return {
        STATUS_REQUESTED: "🟡 Запрошено (ожидает решения директора)",
        STATUS_APPROVED: "🟢 Одобрено (ждёт выдачу)",
        STATUS_REJECTED: "🔴 Отклонено",
        STATUS_ISSUED: "📦 Выдано (ждёт возврат)",
        STATUS_RETURNED: "✅ Возвращено",
    }.get(status, status)


def status_ru(status: str) -> str:
    return {
        STATUS_REQUESTED: "Запрошено",
        STATUS_APPROVED: "Одобрено",
        STATUS_REJECTED: "Отклонено",
        STATUS_ISSUED: "Выдано",
        STATUS_RETURNED: "Возвращено",
    }.get(status, status)


def action_ru(action: str) -> str:
    return {
        "REQUESTED": "Создана заявка",
        "APPROVED": "Одобрено директором",
        "REJECTED": "Отклонено директором",
        "ISSUED": "Выдано уполномоченным",
        "RETURNED": "Возвращено уполномоченным",
    }.get(action, action)


def _format_items(items: Optional[List[Dict[str, Any]]]) -> Tuple[str, str, str]:
    """
    Возвращает:
      companies_str: "ООО Кустос; ООО Поле"
      tokens_str:    "KEY-01; KEY-02"
      items_json:    JSON-строка списка items
    """
    if not items:
        return "", "", ""

    companies: List[str] = []
    tokens: List[str] = []
    norm_items: List[Dict[str, str]] = []

    for it in items:
        c = str(it.get("company", "")).strip()
        t = str(it.get("token_id", "")).strip()
        if c:
            companies.append(c)
        if t:
            tokens.append(t)
        norm_items.append({"company": c, "token_id": t})

    companies_str = "; ".join(companies)
    tokens_str = "; ".join(tokens)
    items_json = json.dumps(norm_items, ensure_ascii=False)
    return companies_str, tokens_str, items_json


def request_card_text(r: Any, items: Optional[List[Dict[str, Any]]] = None) -> str:
    """
    Если items переданы — показываем список "Компания — Токен" (мультизаявка).
    Если items не переданы — показываем одиночную заявку (совместимость).
    """
    d = asdict(r) if hasattr(r, "__dataclass_fields__") else dict(r)

    rid = d.get("id")
    username = d.get("username") or ""
    tg_id = d.get("tg_id")
    company = d.get("company")
    token_id = d.get("token_id")
    purpose = d.get("purpose") or ""
    comment = d.get("comment") or ""
    status = d.get("status") or ""
    requested_at = d.get("requested_at")

    lines = [
        f"<b>Заявка #{rid}</b>",
        f"Статус: <b>{escape_html(status_human(status))}</b>",
    ]

    if requested_at:
        try:
            dt = datetime.fromisoformat(str(requested_at).replace("Z", "+00:00"))
            moscow_tz = ZoneInfo("Europe/Moscow")
            dt_local = dt.astimezone(moscow_tz)
            lines.append(f"📅 Создана: {dt_local.strftime('%d.%m.%Y %H:%M')}")
        except Exception:
            lines.append(f"📅 Создана: {escape_html(str(requested_at))}")

    lines.append("")
    if username:
        lines.append(f"👤 Пользователь: <b>{escape_html(username)}</b> (tg_id: <code>{tg_id}</code>)")
    else:
        lines.append(f"👤 Пользователь: tg_id <code>{tg_id}</code>")

    lines.append(f"🎯 Цель: {escape_html(purpose)}")

    if comment.strip():
        lines.append(f"💬 Комментарий: {escape_html(comment)}")

    if items:
        block = "\n".join(
            f"• <b>{escape_html(str(it.get('company', '')))}</b> — "
            f"<code>{escape_html(str(it.get('token_id', '')))}</code>"
            for it in items
        )
        lines.append("")
        lines.append("<b>Компании / токены:</b>")
        lines.append(block)
    else:
        if company and str(company) != "MULTI":
            lines.append(f"🏢 Компания: <b>{escape_html(str(company))}</b>")
        if token_id and str(token_id) != "MULTI":
            lines.append(f"🔑 Токен: <b>{escape_html(str(token_id))}</b>")

    return "\n".join(lines)


def format_statistics(stats: Dict[str, Any]) -> str:
    lines = ["📊 <b>Статистика системы</b>", ""]

    req = stats.get("requests", {})
    lines.append("<b>Заявки:</b>")
    lines.append(f"  Всего: {req.get('total', 0)}")
    lines.append(f"  На согласовании: {req.get('pending', 0)}")
    lines.append(f"  Одобрено: {req.get('approved', 0)}")
    lines.append(f"  Выдано: {req.get('issued', 0)}")
    lines.append(f"  Возвращено: {req.get('returned', 0)}")
    lines.append(f"  Отклонено: {req.get('rejected', 0)}")

    tokens = stats.get("tokens", {})
    lines.append("")
    lines.append("<b>Токены:</b>")
    for st, count in tokens.items():
        lines.append(f"  {st}: {count}")

    lines.append("")
    lines.append("<b>Пользователи:</b>")
    lines.append(f"  Всего: {stats.get('users_count', 0)}")
    lines.append(f"  Авторизованных: {stats.get('authed_count', 0)}")

    return "\n".join(lines)


# -------------------------
# Nextcloud WebDAV Journal
# -------------------------
MOSCOW_TZ = ZoneInfo("Europe/Moscow")
_journal_lock = asyncio.Lock()


def _msk_now_iso() -> str:
    return datetime.now(MOSCOW_TZ).replace(microsecond=0).isoformat()


def _ensure_sheet(wb) -> Worksheet:
    """
    Поддерживает старый формат файла и новый (мультизаявка).
    Если шапка старая — добавит новые колонки в конец.
    """
    ws = wb.active

    def _is_empty_sheet() -> bool:
        return ws.max_row == 1 and ws.max_column == 1 and ws["A1"].value is None

    header = [
        "ts_msk",
        "request_id",
        "action",
        "actor_tg_id",
        "user_tg_id",
        "username",
        "company",
        "token_id",
        "purpose",
        "comment",
        "status",
        "companies",   # NEW
        "tokens",      # NEW
        "items_json",  # NEW
    ]

    if _is_empty_sheet():
        ws.title = "Journal"
        ws.append(header)
        return ws

    if not ws.title:
        ws.title = "Journal"

    first_row = [ws.cell(row=1, column=c).value for c in range(1, ws.max_column + 1)]
    first_row_str = [str(x).strip() if x is not None else "" for x in first_row]

    if "ts_msk" in first_row_str and "request_id" in first_row_str:
        existing = set(first_row_str)
        col = ws.max_column + 1
        for name in header:
            if name not in existing:
                ws.cell(row=1, column=col).value = name
                col += 1

    return ws


async def append_journal_row(
    *,
    webdav_url: str,
    nc_user: str,
    nc_app_password: str,
    journal_path: str,
    request_row: Any,
    action: str,
    actor_tg_id: int,
    request_items: Optional[List[Dict[str, Any]]] = None,
) -> None:
    async with _journal_lock:
        try:
            await asyncio.to_thread(
                _append_journal_row_sync,
                webdav_url,
                nc_user,
                nc_app_password,
                journal_path,
                request_row,
                action,
                actor_tg_id,
                request_items,
            )
        except Exception as e:
            log.error(f"Failed to append journal row: {e}")


def _append_journal_row_sync(
    webdav_url: str,
    nc_user: str,
    nc_app_password: str,
    journal_path: str,
    request_row: Any,
    action: str,
    actor_tg_id: int,
    request_items: Optional[List[Dict[str, Any]]],
) -> None:
    d = asdict(request_row) if hasattr(request_row, "__dataclass_fields__") else dict(request_row)

    options = {
        "webdav_hostname": webdav_url.rstrip("/") + "/",
        "webdav_login": nc_user,
        "webdav_password": nc_app_password,
        "disable_check": True,
    }
    client = Client(options)

    bio = io.BytesIO()
    exists = False

    try:
        client.download_from(bio, remote_path=journal_path)
        bio.seek(0)
        if bio.getbuffer().nbytes > 0:
            wb = load_workbook(filename=bio)
            exists = True
        else:
            exists = False
    except Exception as e:
        log.warning(f"Could not download journal file: {e}")
        exists = False

    if not exists:
        wb = Workbook()

    ws = _ensure_sheet(wb)

    companies_str, tokens_str, items_json = _format_items(request_items)

    ws.append(
        [
            _msk_now_iso(),
            d.get("id"),
            action_ru(action),
            actor_tg_id,
            d.get("tg_id"),
            d.get("username") or "",
            d.get("company"),
            d.get("token_id"),
            d.get("purpose"),
            d.get("comment") or "",
            status_ru(d.get("status")),
            companies_str,
            tokens_str,
            items_json,
        ]
    )

    out = io.BytesIO()
    wb.save(out)
    out.seek(0)

    try:
        client.upload_to(out, remote_path=journal_path)
    except Exception as e:
        log.error(f"Failed to upload journal: {e}")
        try:
            client.clean(remote_path=journal_path)
            out.seek(0)
            client.upload_to(out, remote_path=journal_path)
        except Exception as e2:
            log.error(f"Failed to recreate journal: {e2}")
            raise


async def webdav_healthcheck(
    webdav_url: str,
    nc_user: str,
    nc_app_password: str,
    journal_path: str,
) -> Tuple[bool, str]:
    try:
        options = {
            "webdav_hostname": webdav_url.rstrip("/") + "/",
            "webdav_login": nc_user,
            "webdav_password": nc_app_password,
            "disable_check": True,
        }
        client = Client(options)
        client.list("/")

        try:
            client.info(remote_path=journal_path)
            return True, "OK (journal exists)"
        except Exception:
            return True, "OK (journal will be created)"
    except Exception as e:
        return False, f"{type(e).__name__}: {str(e)}"


# -------------------------
# Admin helpers (PIN + dashboard)
# -------------------------
def is_director(tg_id: int, settings) -> bool:
    return int(tg_id) == int(getattr(settings, "director_tg_id", -1))


def is_officer(tg_id: int, settings) -> bool:
    return int(tg_id) == int(getattr(settings, "officer_tg_id", -1))


def is_superadmin(tg_id: int, settings) -> bool:
    return int(tg_id) in set(getattr(settings, "superadmin_ids", []) or [])


def kb_admin_menu() -> InlineKeyboardMarkup:
    kb = InlineKeyboardBuilder()
    kb.button(text="📊 Статистика", callback_data="adm:stats")
    kb.button(text="🧑‍💼 На согласовании", callback_data="adm:pending")
    kb.button(text="✅ Одобрено", callback_data="adm:approved")
    kb.button(text="📦 Выдано (на руках)", callback_data="adm:issued")
    kb.button(text="🟢 Активные", callback_data="adm:active")
    kb.button(text="🕒 Последние 20", callback_data="adm:last20")
    kb.button(text="🔑 Все токены", callback_data="adm:tokens")
    kb.button(text="👥 Авторизованные", callback_data="adm:authed")
    kb.button(text="⏱ Висяки >30 мин", callback_data="adm:over:1800")
    kb.button(text="⏱ Висяки >2 часа", callback_data="adm:over:7200")
    kb.button(text="⏱ Висяки >1 день", callback_data="adm:over:86400")
    kb.button(text="🧹 Очистка старых", callback_data="adm:cleanup")
    kb.button(text="🔄 Проверить WebDAV", callback_data="adm:webdav")
    kb.adjust(1, 1, 1, 1, 1, 1, 1, 1, 2, 2, 1, 1)
    return kb.as_markup()


async def safe_edit_text(call: CallbackQuery, text: str, reply_markup: InlineKeyboardMarkup | None = None) -> None:
    try:
        await call.message.edit_text(text, reply_markup=reply_markup)
    except TelegramBadRequest as e:
        if "message is not modified" in str(e):
            try:
                await call.answer()
            except Exception:
                pass
            return
        raise
    except Exception as e:
        log.error(f"Error in safe_edit_text: {e}")
        await call.answer("Ошибка при обновлении сообщения", show_alert=True)


def kb_back_to_admin() -> InlineKeyboardMarkup:
    kb = InlineKeyboardBuilder()
    kb.button(text="⬅️ Назад", callback_data="adm:menu")
    return kb.as_markup()


def format_token_list(tokens: List[Dict[str, Any]]) -> str:
    if not tokens:
        return "Нет токенов в базе данных."

    lines = ["<b>Список токенов:</b>", ""]

    by_status: Dict[str, List[Dict[str, Any]]] = {}
    for token in tokens:
        status = token.get("status", "unknown")
        by_status.setdefault(status, []).append(token)

    for status, token_list in by_status.items():
        status_text = {
            "available": "✅ Доступны",
            "reserved": "🟡 Зарезервированы",
            "issued": "📦 Выданы",
        }.get(status, status)

        lines.append(f"<b>{status_text} ({len(token_list)}):</b>")
        for token in token_list:
            desc = token.get("description", "")
            lines.append(f"  • <code>{token['token_id']}</code> - {escape_html(desc)}")
        lines.append("")

    return "\n".join(lines)
